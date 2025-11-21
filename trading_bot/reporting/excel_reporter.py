"""Excel reporting system for trade analysis and strategy optimization."""

import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import json

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available. Install with: pip install openpyxl")

logger = logging.getLogger(__name__)


class ExcelTradingReporter:
    """Comprehensive Excel reporting for trading analysis and optimization."""
    
    def __init__(self, data_dir: str = "data", reports_dir: str = "reports"):
        """Initialize Excel reporter."""
        self.data_dir = Path(data_dir)
        self.reports_dir = Path(reports_dir)
        self.reports_dir.mkdir(exist_ok=True)
        
        # Load trade data
        self.trades_file = self.data_dir / "daily_trades.json"
        self.trades_data = self._load_trades_data()
        
        logger.info("Excel reporter initialized with %d trades", len(self.trades_data))
    
    def generate_comprehensive_report(self, days_back: int = 30) -> str:
        """Generate comprehensive Excel report with all analysis."""
        if not EXCEL_AVAILABLE:
            logger.error("Cannot generate Excel report - openpyxl not installed")
            return ""
        
        try:
            # Create filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Trading_Analysis_Report_{timestamp}.xlsx"
            filepath = self.reports_dir / filename
            
            # Filter trades for analysis period
            cutoff_date = datetime.now() - timedelta(days=days_back)
            filtered_trades = self._filter_trades_by_date(cutoff_date)
            
            if not filtered_trades:
                logger.warning("No trades found in the last %d days", days_back)
                return ""
            
            # Create Excel workbook
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Trade Details
                self._create_trades_sheet(writer, filtered_trades)
                
                # Sheet 2: Performance Summary
                self._create_performance_sheet(writer, filtered_trades)
                
                # Sheet 3: Strategy Analysis
                self._create_strategy_sheet(writer, filtered_trades)
                
                # Sheet 4: Risk Analysis
                self._create_risk_sheet(writer, filtered_trades)
                
                # Sheet 5: Symbol Performance
                self._create_symbol_sheet(writer, filtered_trades)
                
                # Sheet 6: Time Analysis
                self._create_time_sheet(writer, filtered_trades)
                
                # Sheet 7: Recommendations
                self._create_recommendations_sheet(writer, filtered_trades)
            
            # Apply formatting and charts
            self._apply_formatting_and_charts(filepath)
            
            logger.info("✅ Excel report generated: %s", filepath)
            return str(filepath)
            
        except Exception as exc:
            logger.error("Failed to generate Excel report: %s", exc)
            return ""
    
    def _load_trades_data(self) -> List[Dict]:
        """Load trade data from JSON file."""
        try:
            if self.trades_file.exists():
                with open(self.trades_file, 'r') as f:
                    return json.load(f)
            return []
        except Exception as exc:
            logger.error("Error loading trades data: %s", exc)
            return []
    
    def _filter_trades_by_date(self, cutoff_date: datetime) -> List[Dict]:
        """Filter trades by date range."""
        filtered = []
        for trade in self.trades_data:
            if trade.get('exit_time'):
                trade_date = datetime.fromtimestamp(trade['exit_time'])
                if trade_date >= cutoff_date:
                    filtered.append(trade)
        return filtered
    
    def _create_trades_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create detailed trades sheet."""
        # Convert trades to DataFrame
        df_data = []
        for trade in trades:
            entry_time = datetime.fromtimestamp(trade['entry_time']) if trade.get('entry_time') else None
            exit_time = datetime.fromtimestamp(trade['exit_time']) if trade.get('exit_time') else None
            
            duration_hours = 0
            if entry_time and exit_time:
                duration_hours = (exit_time - entry_time).total_seconds() / 3600
            
            df_data.append({
                'Trade_ID': f"{trade['symbol']}_{int(trade.get('entry_time', 0))}",
                'Symbol': trade['symbol'],
                'Entry_Date': entry_time.strftime('%Y-%m-%d') if entry_time else '',
                'Entry_Time': entry_time.strftime('%H:%M:%S') if entry_time else '',
                'Exit_Date': exit_time.strftime('%Y-%m-%d') if exit_time else '',
                'Exit_Time': exit_time.strftime('%H:%M:%S') if exit_time else '',
                'Duration_Hours': round(duration_hours, 2),
                'Entry_Price': trade.get('entry_price', 0),
                'Exit_Price': trade.get('exit_price', 0),
                'Amount': trade.get('amount', 0),
                'PnL_USD': trade.get('pnl_usd', 0),
                'PnL_Percentage': trade.get('pnl_percentage', 0),
                'Confidence': trade.get('confidence', 0),
                'Exit_Reason': trade.get('reason', ''),
                'Win_Loss': 'WIN' if trade.get('pnl_usd', 0) > 0 else 'LOSS',
                'Trade_Size_USD': trade.get('entry_price', 0) * trade.get('amount', 0)
            })
        
        df = pd.DataFrame(df_data)
        df.to_excel(writer, sheet_name='Trade_Details', index=False)
        
        logger.info("Created Trade Details sheet with %d trades", len(df))
    
    def _create_performance_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create performance summary sheet."""
        # Calculate KPIs
        total_trades = len(trades)
        winning_trades = len([t for t in trades if t.get('pnl_usd', 0) > 0])
        losing_trades = total_trades - winning_trades
        
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        
        total_pnl = sum(t.get('pnl_usd', 0) for t in trades)
        total_pnl_pct = sum(t.get('pnl_percentage', 0) for t in trades)
        
        wins = [t.get('pnl_usd', 0) for t in trades if t.get('pnl_usd', 0) > 0]
        losses = [abs(t.get('pnl_usd', 0)) for t in trades if t.get('pnl_usd', 0) < 0]
        
        avg_win = np.mean(wins) if wins else 0
        avg_loss = np.mean(losses) if losses else 0
        
        profit_factor = sum(wins) / sum(losses) if losses else float('inf')
        
        # Calculate Sharpe ratio (simplified)
        returns = [t.get('pnl_percentage', 0) for t in trades]
        sharpe_ratio = np.mean(returns) / np.std(returns) if len(returns) > 1 and np.std(returns) > 0 else 0
        
        # Calculate maximum drawdown
        cumulative_returns = np.cumsum([t.get('pnl_usd', 0) for t in trades])
        running_max = np.maximum.accumulate(cumulative_returns)
        drawdown = (cumulative_returns - running_max)
        max_drawdown = np.min(drawdown) if len(drawdown) > 0 else 0
        
        # Create performance summary
        performance_data = {
            'KPI': [
                'Total Trades', 'Winning Trades', 'Losing Trades', 'Win Rate (%)',
                'Total PnL (USD)', 'Total PnL (%)', 'Average Win (USD)', 'Average Loss (USD)',
                'Profit Factor', 'Sharpe Ratio', 'Max Drawdown (USD)', 'Best Trade (USD)',
                'Worst Trade (USD)', 'Average Trade Duration (Hours)', 'Average Confidence'
            ],
            'Value': [
                total_trades, winning_trades, losing_trades, round(win_rate, 2),
                round(total_pnl, 2), round(total_pnl_pct, 2), round(avg_win, 2), round(avg_loss, 2),
                round(profit_factor, 2), round(sharpe_ratio, 2), round(max_drawdown, 2),
                round(max(wins) if wins else 0, 2), round(-min([t.get('pnl_usd', 0) for t in trades]), 2),
                round(np.mean([self._calculate_duration(t) for t in trades]), 2),
                round(np.mean([t.get('confidence', 0) for t in trades]), 2)
            ],
            'Target': [
                '>50', '>30', '<20', '>60%',
                '>0', '>0', '>50', '<30',
                '>1.5', '>1.0', '<-200', '', '',
                '<24', '>0.6'
            ],
            'Status': []
        }
        
        # Add status based on targets
        targets = [50, 30, 20, 60, 0, 0, 50, 30, 1.5, 1.0, -200, 0, 0, 24, 0.6]
        comparisons = ['>=', '>=', '<=', '>=', '>=', '>=', '>=', '<=', '>=', '>=', '>=', '', '', '<=', '>=']
        
        for i, (value, target, comp) in enumerate(zip(performance_data['Value'], targets, comparisons)):
            if comp == '>=':
                status = '✅ GOOD' if value >= target else '❌ POOR'
            elif comp == '<=':
                status = '✅ GOOD' if value <= target else '❌ POOR'
            else:
                status = '➖ N/A'
            performance_data['Status'].append(status)
        
        df = pd.DataFrame(performance_data)
        df.to_excel(writer, sheet_name='Performance_Summary', index=False)
        
        logger.info("Created Performance Summary sheet")
    
    def _create_strategy_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create strategy analysis sheet."""
        # Analyze by exit reason
        exit_reasons = {}
        for trade in trades:
            reason = trade.get('reason', 'unknown')
            if reason not in exit_reasons:
                exit_reasons[reason] = {'count': 0, 'total_pnl': 0, 'wins': 0}
            
            exit_reasons[reason]['count'] += 1
            exit_reasons[reason]['total_pnl'] += trade.get('pnl_usd', 0)
            if trade.get('pnl_usd', 0) > 0:
                exit_reasons[reason]['wins'] += 1
        
        # Create strategy analysis DataFrame
        strategy_data = []
        for reason, stats in exit_reasons.items():
            win_rate = (stats['wins'] / stats['count'] * 100) if stats['count'] > 0 else 0
            avg_pnl = stats['total_pnl'] / stats['count'] if stats['count'] > 0 else 0
            
            strategy_data.append({
                'Exit_Reason': reason,
                'Trade_Count': stats['count'],
                'Win_Count': stats['wins'],
                'Win_Rate_%': round(win_rate, 2),
                'Total_PnL_USD': round(stats['total_pnl'], 2),
                'Avg_PnL_USD': round(avg_pnl, 2),
                'Effectiveness': 'HIGH' if win_rate > 70 else 'MEDIUM' if win_rate > 50 else 'LOW'
            })
        
        df = pd.DataFrame(strategy_data)
        df = df.sort_values('Total_PnL_USD', ascending=False)
        df.to_excel(writer, sheet_name='Strategy_Analysis', index=False)
        
        logger.info("Created Strategy Analysis sheet")
    
    def _create_risk_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create risk analysis sheet."""
        # Risk metrics by trade size
        trade_sizes = []
        for trade in trades:
            size_usd = trade.get('entry_price', 0) * trade.get('amount', 0)
            trade_sizes.append({
                'Trade_Size_USD': size_usd,
                'PnL_USD': trade.get('pnl_usd', 0),
                'PnL_Percentage': trade.get('pnl_percentage', 0),
                'Duration_Hours': self._calculate_duration(trade),
                'Confidence': trade.get('confidence', 0)
            })
        
        df = pd.DataFrame(trade_sizes)
        
        # Create size buckets
        df['Size_Bucket'] = pd.cut(df['Trade_Size_USD'], 
                                  bins=[0, 100, 500, 1000, 5000, float('inf')],
                                  labels=['<$100', '$100-500', '$500-1K', '$1K-5K', '>$5K'])
        
        # Risk analysis by bucket
        risk_analysis = df.groupby('Size_Bucket').agg({
            'PnL_USD': ['count', 'mean', 'sum', 'std'],
            'PnL_Percentage': ['mean', 'std'],
            'Confidence': 'mean'
        }).round(2)
        
        risk_analysis.columns = ['Trade_Count', 'Avg_PnL_USD', 'Total_PnL_USD', 'PnL_Volatility',
                                'Avg_PnL_Pct', 'Pct_Volatility', 'Avg_Confidence']
        
        risk_analysis.to_excel(writer, sheet_name='Risk_Analysis')
        
        logger.info("Created Risk Analysis sheet")
    
    def _create_symbol_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create symbol performance analysis."""
        # Analyze by symbol
        symbols = {}
        for trade in trades:
            symbol = trade['symbol']
            if symbol not in symbols:
                symbols[symbol] = {
                    'trades': [], 'total_pnl': 0, 'wins': 0, 'total_confidence': 0
                }
            
            symbols[symbol]['trades'].append(trade)
            symbols[symbol]['total_pnl'] += trade.get('pnl_usd', 0)
            symbols[symbol]['total_confidence'] += trade.get('confidence', 0)
            if trade.get('pnl_usd', 0) > 0:
                symbols[symbol]['wins'] += 1
        
        # Create symbol analysis
        symbol_data = []
        for symbol, stats in symbols.items():
            trade_count = len(stats['trades'])
            win_rate = (stats['wins'] / trade_count * 100) if trade_count > 0 else 0
            avg_confidence = stats['total_confidence'] / trade_count if trade_count > 0 else 0
            
            # Calculate volatility
            pnl_values = [t.get('pnl_percentage', 0) for t in stats['trades']]
            volatility = np.std(pnl_values) if len(pnl_values) > 1 else 0
            
            symbol_data.append({
                'Symbol': symbol,
                'Trade_Count': trade_count,
                'Win_Count': stats['wins'],
                'Win_Rate_%': round(win_rate, 2),
                'Total_PnL_USD': round(stats['total_pnl'], 2),
                'Avg_PnL_USD': round(stats['total_pnl'] / trade_count, 2),
                'Avg_Confidence': round(avg_confidence, 2),
                'Volatility_%': round(volatility, 2),
                'Performance': 'EXCELLENT' if win_rate > 80 else 'GOOD' if win_rate > 60 else 'POOR'
            })
        
        df = pd.DataFrame(symbol_data)
        df = df.sort_values('Total_PnL_USD', ascending=False)
        df.to_excel(writer, sheet_name='Symbol_Performance', index=False)
        
        logger.info("Created Symbol Performance sheet")
    
    def _create_time_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create time-based analysis."""
        # Analyze by hour of day and day of week
        time_data = []
        for trade in trades:
            if trade.get('entry_time'):
                entry_dt = datetime.fromtimestamp(trade['entry_time'])
                time_data.append({
                    'Hour': entry_dt.hour,
                    'Day_of_Week': entry_dt.strftime('%A'),
                    'Date': entry_dt.date(),
                    'PnL_USD': trade.get('pnl_usd', 0),
                    'Win': 1 if trade.get('pnl_usd', 0) > 0 else 0
                })
        
        df = pd.DataFrame(time_data)
        
        # Hourly analysis
        hourly = df.groupby('Hour').agg({
            'PnL_USD': ['count', 'sum', 'mean'],
            'Win': 'sum'
        }).round(2)
        hourly.columns = ['Trade_Count', 'Total_PnL', 'Avg_PnL', 'Wins']
        hourly['Win_Rate_%'] = (hourly['Wins'] / hourly['Trade_Count'] * 100).round(2)
        
        # Daily analysis
        daily = df.groupby('Day_of_Week').agg({
            'PnL_USD': ['count', 'sum', 'mean'],
            'Win': 'sum'
        }).round(2)
        daily.columns = ['Trade_Count', 'Total_PnL', 'Avg_PnL', 'Wins']
        daily['Win_Rate_%'] = (daily['Wins'] / daily['Trade_Count'] * 100).round(2)
        
        # Write to Excel with multiple tables
        hourly.to_excel(writer, sheet_name='Time_Analysis', startrow=0)
        daily.to_excel(writer, sheet_name='Time_Analysis', startrow=len(hourly) + 3)
        
        logger.info("Created Time Analysis sheet")
    
    def _create_recommendations_sheet(self, writer: pd.ExcelWriter, trades: List[Dict]) -> None:
        """Create recommendations based on analysis."""
        recommendations = []
        
        # Analyze performance and generate recommendations
        total_trades = len(trades)
        winning_trades = len([t for t in trades if t.get('pnl_usd', 0) > 0])
        win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
        
        # Win rate recommendations
        if win_rate < 50:
            recommendations.append({
                'Category': 'Win Rate',
                'Issue': f'Low win rate: {win_rate:.1f}%',
                'Recommendation': 'Increase confidence threshold, improve entry signals',
                'Priority': 'HIGH',
                'Expected_Impact': 'Increase win rate by 10-15%'
            })
        
        # Analyze exit reasons
        exit_reasons = {}
        for trade in trades:
            reason = trade.get('reason', 'unknown')
            if reason not in exit_reasons:
                exit_reasons[reason] = {'count': 0, 'pnl': 0}
            exit_reasons[reason]['count'] += 1
            exit_reasons[reason]['pnl'] += trade.get('pnl_usd', 0)
        
        # Find worst performing exit reasons
        worst_reasons = sorted(exit_reasons.items(), key=lambda x: x[1]['pnl'])[:3]
        for reason, stats in worst_reasons:
            if stats['pnl'] < -50:  # Significant losses
                recommendations.append({
                    'Category': 'Exit Strategy',
                    'Issue': f'Poor performance from {reason}: ${stats["pnl"]:.2f}',
                    'Recommendation': f'Review and optimize {reason} exit conditions',
                    'Priority': 'MEDIUM',
                    'Expected_Impact': 'Reduce losses by 20-30%'
                })
        
        # Position sizing recommendations
        trade_sizes = [t.get('entry_price', 0) * t.get('amount', 0) for t in trades]
        if trade_sizes:
            avg_size = np.mean(trade_sizes)
            size_std = np.std(trade_sizes)
            if size_std / avg_size > 0.5:  # High variability
                recommendations.append({
                    'Category': 'Position Sizing',
                    'Issue': 'Inconsistent position sizing detected',
                    'Recommendation': 'Implement fixed percentage position sizing (5% recommended)',
                    'Priority': 'HIGH',
                    'Expected_Impact': 'Improve risk management and consistency'
                })
        
        # Confidence analysis
        confidences = [t.get('confidence', 0) for t in trades if t.get('confidence', 0) > 0]
        if confidences:
            avg_confidence = np.mean(confidences)
            if avg_confidence < 0.6:
                recommendations.append({
                    'Category': 'Signal Quality',
                    'Issue': f'Low average confidence: {avg_confidence:.2f}',
                    'Recommendation': 'Increase minimum confidence threshold to 0.6+',
                    'Priority': 'MEDIUM',
                    'Expected_Impact': 'Higher quality trades, better win rate'
                })
        
        # Add general recommendations
        recommendations.extend([
            {
                'Category': 'Risk Management',
                'Issue': 'Continuous improvement needed',
                'Recommendation': 'Implement trailing stops for profitable positions',
                'Priority': 'LOW',
                'Expected_Impact': 'Protect profits, reduce drawdowns'
            },
            {
                'Category': 'Diversification',
                'Issue': 'Portfolio optimization',
                'Recommendation': 'Limit maximum position size to 15% of portfolio',
                'Priority': 'MEDIUM',
                'Expected_Impact': 'Reduce concentration risk'
            }
        ])
        
        df = pd.DataFrame(recommendations)
        df.to_excel(writer, sheet_name='Recommendations', index=False)
        
        logger.info("Created Recommendations sheet with %d suggestions", len(recommendations))
    
    def _calculate_duration(self, trade: Dict) -> float:
        """Calculate trade duration in hours."""
        if trade.get('entry_time') and trade.get('exit_time'):
            return (trade['exit_time'] - trade['entry_time']) / 3600
        return 0
    
    def _apply_formatting_and_charts(self, filepath: Path) -> None:
        """Apply Excel formatting and add charts."""
        try:
            wb = openpyxl.load_workbook(filepath)
            
            # Format Performance Summary sheet
            if 'Performance_Summary' in wb.sheetnames:
                ws = wb['Performance_Summary']
                self._format_performance_sheet(ws)
            
            # Format other sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._apply_basic_formatting(ws)
            
            wb.save(filepath)
            logger.info("Applied Excel formatting and charts")
            
        except Exception as exc:
            logger.error("Error applying Excel formatting: %s", exc)
    
    def _format_performance_sheet(self, ws) -> None:
        """Format the performance summary sheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Status column formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                if "✅" in str(cell.value):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "❌" in str(cell.value):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _apply_basic_formatting(self, ws) -> None:
        """Apply basic formatting to worksheet."""
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border


def generate_trading_report(days_back: int = 30) -> str:
    """Generate comprehensive trading report."""
    reporter = ExcelTradingReporter()
    return reporter.generate_comprehensive_report(days_back)


if __name__ == "__main__":
    # Generate sample report
    report_path = generate_trading_report(30)
    if report_path:
        print(f"Report generated: {report_path}")
    else:
        print("Failed to generate report")
